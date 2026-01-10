from openpyxl import Workbook

from django.db.models import Q
from django.http import HttpResponse
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum
from datetime import datetime

from apps.organizations.models import Organization
from apps.organizations.serializers import OrganizationSerializer
from apps.products.models import Product
from apps.transactions.models import Transaction
from apps.transactions.serializers import GetTransactionSerializer


class OrganizationViewSet(ModelViewSet):
    serializer_class = OrganizationSerializer
    queryset = Organization.objects.all()

    @action(methods=["GET"], detail=True, serializer_class=GetTransactionSerializer)
    def transactions(self, request, *args, **kwargs):
        obj: Organization = self.get_object()

        queryset = (
            Transaction.objects.select_related("sender", "receiver", "project")
            .prefetch_related(
                "items",
                "items__product__material",
                "items__product__organization",
                "items__product__project",
            )
            .filter(Q(sender=obj) | Q(receiver=obj))
            .order_by("-id")
        )

        start_date = request.GET.get("start_date", None)
        end_date = request.GET.get("end_date", None)

        if start_date and end_date:
            queryset = queryset.filter(created_at__range=[start_date, end_date])

        elif start_date:
            queryset = queryset.filter(created_at__gte=start_date)

        elif end_date:
            queryset = queryset.filter(created_at__lte=end_date)

        serializer = self.get_serializer(queryset, many=True)

        return Response(serializer.data)

    @action(methods=["POST"], detail=False)
    def report(self, request, *args, **kwargs):
        wb = Workbook()

        # Default worksheet'ni o'chiramiz
        wb.remove(wb.active)

        # Har bir organization uchun alohida sheet yaratamiz
        organizations = Organization.objects.all()

        for org in organizations:
            # Organization uchun products filter qilamiz
            org_products = Product.objects.filter(organization=org)

            # Sheet yaratamiz
            ws = wb.create_sheet(title=f"{org.name}")

            # Header qo'shamiz
            ws.append(["ID", "Material", "Miqdori (Quantity)", "Proba (Purity)", "Sof oltin (Pure Gold)", "Kompozit (Is Composite)", "Manba tavsifi (Source Description)", "Yaratilgan sana (Created At)"])

            # Har bir product uchun qator qo'shamiz
            for product in org_products:
                ws.append([product.id, product.material.name, float(product.quantity), float(product.purity), float(product.pure_gold), "Ha" if product.is_composite else "Yo'q", product.source_description or "", product.created_at.strftime("%Y-%m-%d %H:%M:%S")])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=products_by_organizations.xlsx"

        wb.save(response)
        return response

    @action(methods=["POST"], detail=True)
    def organization_report(self, request, *args, **kwargs):
        """Muayyan organizatsiya uchun individual report"""
        organization: Organization = self.get_object()

        # Organization products
        org_products = Product.objects.filter(organization=organization).select_related("material", "project")

        wb = Workbook()
        ws = wb.active
        ws.title = f"{organization.name}"

        # Header
        ws.append(["ID", "Material", "Proyekt", "Miqdori (Quantity)", "Proba (Purity)", "Sof oltin (Pure Gold)", "Kompozit (Is Composite)", "Manba tavsifi (Source Description)", "Yaratilgan sana (Created At)"])

        # Data rows
        for product in org_products:
            ws.append([product.id, product.material.name, product.project.name if product.project else "", float(product.quantity), float(product.purity), float(product.pure_gold), "Ha" if product.is_composite else "Yo'q", product.source_description or "", product.created_at.strftime("%Y-%m-%d %H:%M:%S")])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={organization.name}_products.xlsx"

        wb.save(response)
        return response

    @action(methods=["POST"], detail=False, url_path="products-count-matrix")
    def products_count_matrix(self, request, *args, **kwargs):
        """
        Mahsulotlar sonini matrix shaklida ko'rsatish.
        Birinchi qator: Organizatsiyalar nomi
        Birinchi ustun: Mahsulotlar nomi
        Celllar: Mahsulot soni (quantity)
        """

        # Barcha organizatsiyalarni olish
        organizations = Organization.objects.all().order_by("name")

        # Barcha materiallarni olish (products'ni group by material)
        materials = Product.objects.values("material__name", "material__purity").annotate(total=Sum("quantity")).order_by("material__name")

        # Material nomlari
        material_names = [f"{m['material__name']} ({round(m['material__purity'], 2)})" for m in materials]

        # Organization nomlari
        org_names = list(organizations.values_list("name", flat=True))

        wb = Workbook()
        ws = wb.active
        ws.title = "Products Count"

        # First row: Organization names
        ws.append(["Material"] + org_names)

        # Data rows
        for material_name in material_names:
            row = [material_name]

            for org in organizations:
                # Har bir organization uchun material miqdorini hisoblash
                count = (Product.objects.filter(organization=org, material__name=material_name).aggregate(total=Sum("quantity"))["total"]) or 0
                row.append(float(count) if count else 0)

            ws.append(row)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        response["Content-Disposition"] = f"attachment; filename=products-count-matrix_{current_datetime}.xlsx"

        wb.save(response)
        return response
